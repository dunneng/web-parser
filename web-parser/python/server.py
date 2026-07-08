"""
网页源码解析器 — Python 后端服务
FastAPI + 解析引擎
"""
import sys
import os
import json
import re
import time
import logging
from pathlib import Path
from contextlib import asynccontextmanager

# 强制 UTF-8 输出，解决 Windows GBK 乱码
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except Exception:
        pass

from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Body
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional
import io
import csv
import sqlite3
import openpyxl

# 确保能找到 parser 模块
sys.path.insert(0, os.path.dirname(__file__))

from parser.html_parser import format_html, get_source_stats
from parser.dom_parser import build_dom_tree
from parser.script_parser import extract_scripts, parse_json_from_text
from parser.xpath_engine import xpath_query
from parser.css_engine import css_query
from parser.regex_engine import regex_search
from parser.jsonpath_engine import jsonpath_query
from parser.chain_engine import chain_extract, trace_chain_backend
import db
import product_pipeline

logging.basicConfig(
    level=logging.WARNING,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)

COOKIES_DIR = Path(__file__).parent / "cookies"
COOKIES_DIR.mkdir(exist_ok=True)

def _safe_path_component(name: str) -> str:
    """净化路径组件，防止路径遍历攻击"""
    # 只保留字母数字、中划线、下划线和点
    safe = re.sub(r"[^a-zA-Z0-9._-]", "_", name)
    # 防止以点开头（隐藏文件）
    safe = safe.lstrip(".")
    # 空字符串则用默认值
    return safe or "unknown"

@asynccontextmanager
async def lifespan(app: FastAPI):
    db.init_db()
    try:
        product_pipeline.init_pipeline()
        logger.info("解析引擎启动完毕 (SQLite 持久化 + 向量库)")
    except Exception as e:
        logger.warning(f"向量库初始化失败（不影响基本解析功能）: {e}")
    yield

import threading
import uuid

_import_jobs = {}  # {job_id: {status: "running"|"done"|"error", progress: {done, total}, result: {...}}}

app = FastAPI(title="Web Parser Engine", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"]
)

# ── 静态文件 ──
from fastapi.staticfiles import StaticFiles
IMAGES_DIR = os.path.join(os.path.dirname(__file__), "data", "images")
if os.path.isdir(IMAGES_DIR):
    app.mount("/static/images", StaticFiles(directory=IMAGES_DIR), name="static_images")

# ── 比价控制台页面 ──
@app.get("/price-compare")
@app.get("/price-compare/")
async def price_compare_page():
    """返回比价控制台 HTML"""
    ui_path = os.path.join(os.path.dirname(__file__), "price_compare_ui.html")
    from fastapi.responses import HTMLResponse
    with open(ui_path, "r", encoding="utf-8") as f:
        return HTMLResponse(
            content=f.read(),
            headers={
                "Cache-Control": "no-cache, no-store, must-revalidate",
                "Pragma": "no-cache",
                "Expires": "0",
            },
        )

# ── 商品图片服务（比 static mount 更可靠）──
from fastapi.responses import FileResponse, Response
import mimetypes

@app.get("/api/price-compare/image/{product_id}")
async def get_product_image(product_id: int):
    """根据商品 ID 返回本地图片。比 static mount 更可靠，不受挂载失败影响。"""
    product = db.get_product_by_id(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="商品不存在")

    local_image = product.get("local_image", "")
    if local_image and os.path.isfile(local_image):
        mime, _ = mimetypes.guess_type(local_image)
        return FileResponse(local_image, media_type=mime or "image/jpeg")

    # 无本地图片：代理远程主图
    remote = product.get("main_image_url", "")
    if remote:
        return _proxy_remote_image(remote)
    raise HTTPException(status_code=404, detail="无图片")

# ── 多图服务：按索引返回本地图片 ──
@app.get("/api/price-compare/image/{product_id}/{index}")
async def get_product_image_by_index(product_id: int, index: int):
    """根据商品 ID + 图片索引返回本地图片。index=0→主图, index>=1→附图"""
    product = db.get_product_by_id(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="商品不存在")

    platform = product.get("platform", "unknown")
    local_path = _resolve_image_path(platform, product_id, index)

    if local_path and os.path.isfile(local_path):
        mime, _ = mimetypes.guess_type(local_path)
        return FileResponse(local_path, media_type=mime or "image/jpeg")

    # 无本地图片：尝试返回远程 URL（重定向）
    image_urls = product.get("image_urls") or []
    if index == 0:
        remote = product.get("main_image_url", "")
    elif 0 <= (image_idx := index - 1) < len(image_urls):
        remote = image_urls[image_idx]
    else:
        remote = ""

    if remote:
        return _proxy_remote_image(remote)

    raise HTTPException(status_code=404, detail="图片不存在")

def _resolve_image_path(platform: str, product_id: int, index: int) -> str | None:
    """根据命名规则反推本地图片路径"""
    images_dir = os.path.join(os.path.dirname(__file__), "data", "images")
    base = Path(images_dir) / platform
    if index == 0:
        return str(base / f"{product_id}.jpg")
    else:
        return str(base / f"{product_id}_{index}.jpg")

def _proxy_remote_image(url: str) -> Response:
    """代理下载远程图片，复用 proxy-image 缓存逻辑"""
    cache_key = hashlib.sha256(url.encode()).hexdigest()
    with _proxy_cache_lock:
        if cache_key in _proxy_cache:
            data, mime, expires = _proxy_cache[cache_key]
            if _time.time() < expires:
                return Response(content=data, media_type=mime)
    try:
        resp = requests.get(url, timeout=15, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept": "image/avif,image/webp,image/*,*/*;q=0.8",
            "Referer": "",
        })
        resp.raise_for_status()
        content_type = resp.headers.get("Content-Type", "image/jpeg")
        data = resp.content
        if len(data) > PROXY_MAX_SIZE:
            raise HTTPException(status_code=413, detail="图片过大")
        with _proxy_cache_lock:
            _proxy_cache[cache_key] = (data, content_type, _time.time() + PROXY_CACHE_TTL)
        return Response(content=data, media_type=content_type)
    except requests.RequestException as e:
        raise HTTPException(status_code=502, detail=f"下载失败: {str(e)[:100]}")

# ── 图片 CRUD + 查询 ──

class ImageListBody(BaseModel):
    main_image_url: str = ""
    image_urls: list[str] = []

class AddImagesBody(BaseModel):
    image_urls: list[str]

class ReorderBody(BaseModel):
    """重排：new_order[idx] = 新位置（0=主图, 1..N=附图）。"""
    new_order: list[int]

@app.get("/api/price-compare/products/{product_id}/images")
async def get_product_images(product_id: int):
    """返回商品的所有图片信息（本地路径 + 远程 URL）"""
    p = db.get_product_by_id(product_id)
    if not p:
        raise HTTPException(status_code=404, detail="商品不存在")

    platform = p.get("platform", "unknown")
    image_urls = p.get("image_urls") or []
    all_images = []

    # 主图 (index 0)
    local0 = _resolve_image_path(platform, product_id, 0)
    all_images.append({
        "index": 0,
        "url": p.get("main_image_url", ""),
        "local": local0 if (local0 and os.path.isfile(local0)) else "",
        "is_main": True,
    })

    # 附图 (index 1+)
    for i, url in enumerate(image_urls):
        idx = i + 1
        local = _resolve_image_path(platform, product_id, idx)
        all_images.append({
            "index": idx,
            "url": url,
            "local": local if (local and os.path.isfile(local)) else "",
            "is_main": False,
        })

    return {"ok": True, "images": all_images, "count": len(all_images)}

@app.put("/api/price-compare/products/{product_id}/images")
async def replace_product_images(product_id: int, body: ImageListBody):
    """整体替换图片列表（主图 + 附图）"""
    p = db.get_product_by_id(product_id)
    if not p:
        raise HTTPException(status_code=404, detail="商品不存在")
    db.update_product(product_id, main_image_url=body.main_image_url,
                      image_urls=body.image_urls)
    return {"ok": True}

@app.post("/api/price-compare/products/{product_id}/images")
async def add_product_images(product_id: int, body: AddImagesBody):
    """追加图片 URL 到附图列表"""
    p = db.get_product_by_id(product_id)
    if not p:
        raise HTTPException(status_code=404, detail="商品不存在")
    existing = list(p.get("image_urls") or [])
    existing.extend(body.image_urls)
    db.update_product(product_id, image_urls=existing)
    return {"ok": True, "count": len(existing)}

@app.delete("/api/price-compare/products/{product_id}/images/{index}")
async def delete_product_image(product_id: int, index: int):
    """删除指定索引的图片。index=0→清空主图, index>=1→从附图列表移除"""
    p = db.get_product_by_id(product_id)
    if not p:
        raise HTTPException(status_code=404, detail="商品不存在")

    if index == 0:
        db.update_product(product_id, main_image_url="")
        # 同时删除本地文件，否则 /image/{id}/0 会一直返回旧图
        platform = p.get("platform", "unknown")
        local = _resolve_image_path(platform, product_id, 0)
        if local and os.path.isfile(local):
            try:
                os.remove(local)
            except Exception:
                pass
        return {"ok": True}

    image_urls = list(p.get("image_urls") or [])
    real_idx = index - 1
    if real_idx < 0 or real_idx >= len(image_urls):
        raise HTTPException(status_code=404, detail="索引超出范围")
    removed = image_urls.pop(real_idx)
    db.update_product(product_id, image_urls=image_urls)
    # 删除本地文件
    platform = p.get("platform", "unknown")
    local = _resolve_image_path(platform, product_id, index)
    if local and os.path.isfile(local):
        try:
            os.remove(local)
        except Exception:
            pass
    return {"ok": True, "removed": removed, "remaining": len(image_urls)}

@app.put("/api/price-compare/products/{product_id}/images/reorder")
async def reorder_product_images(product_id: int, body: ReorderBody):
    """重排图片。new_order[idx] = 新位置（0=主图, 1..N=附图含空位）"""
    p = db.get_product_by_id(product_id)
    if not p:
        raise HTTPException(status_code=404, detail="商品不存在")

    # 构建有序列表：[(index, url), ...]
    main_url = p.get("main_image_url", "")
    image_urls = list(p.get("image_urls") or [])
    all_urls = [main_url] + image_urls  # index 0=N, 1=first extra, ...
    total = len(all_urls)

    if sorted(body.new_order) != list(range(total)):
        raise HTTPException(400, f"new_order 必须是 0..{total-1} 的排列")

    # 按新顺序重排
    reordered = [all_urls[old_pos] for old_pos in body.new_order]
    new_main = reordered[0] if reordered else ""
    new_extras = reordered[1:] if len(reordered) > 1 else []

    db.update_product(product_id, main_image_url=new_main, image_urls=new_extras)
    return {"ok": True}

# ── 图片代理 ──
import requests
import hashlib
import tempfile
import threading
import time as _time

_proxy_cache: dict[str, tuple[bytes, str, float]] = {}  # cache_key → (bytes, mime, expires_at)
_proxy_cache_lock = threading.Lock()
PROXY_CACHE_TTL = 600  # 10 分钟
PROXY_MAX_SIZE = 10 * 1024 * 1024  # 10MB

def _proxy_cleanup():
    """清理过期缓存"""
    with _proxy_cache_lock:
        now = _time.time()
        expired = [k for k, v in _proxy_cache.items() if v[2] < now]
        for k in expired:
            del _proxy_cache[k]

@app.get("/api/price-compare/proxy-image")
async def proxy_image(url: str):
    """代理下载远程图片"""
    if not url or not (url.startswith("http://") or url.startswith("https://")):
        raise HTTPException(status_code=400, detail="无效的图片 URL")

    cache_key = hashlib.sha256(url.encode()).hexdigest()

    # 检查缓存
    with _proxy_cache_lock:
        if cache_key in _proxy_cache:
            data, mime, expires = _proxy_cache[cache_key]
            if _time.time() < expires:
                return Response(content=data, media_type=mime)

    # 下载
    try:
        resp = requests.get(url, timeout=15, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept": "image/avif,image/webp,image/*,*/*;q=0.8",
            "Referer": "",
        })
        resp.raise_for_status()
        content_type = resp.headers.get("Content-Type", "image/jpeg")
        data = resp.content
        if len(data) > PROXY_MAX_SIZE:
            raise HTTPException(status_code=413, detail="图片过大（>10MB）")

        # 缓存
        with _proxy_cache_lock:
            _proxy_cache[cache_key] = (data, content_type, _time.time() + PROXY_CACHE_TTL)
            # 定期清理（每 60 秒一次）
            if len(_proxy_cache) > 50:
                _proxy_cleanup()

        return Response(content=data, media_type=content_type)
    except requests.RequestException as e:
        raise HTTPException(status_code=502, detail=f"下载失败: {str(e)[:100]}")

# ── 原解析器路由 ──
class ParseRequest(BaseModel):
    html: str
    query: str = ""
    child_delim: str = ""
    max_text_len: int = 2000
    max_depth: int = 20
    max_results: int = 1000
    max_children: int = 200
    expand_children: bool = False

class ChainRequest(BaseModel):
    html: str
    chain_type: str = "css"       # 'css' | 'xpath'
    deepest_selector: str = ""    # 最深层选择器
    fields: list[dict] = []       # [{chainIndex, attr, isText, subChain, name}]
    child_delim: str = ""         # 子节点分隔符

class CookieItem(BaseModel):
    name: str
    value: str
    domain: str
    path: str = "/"
    secure: bool = False
    httpOnly: bool = False
    expirationDate: float = 0
    sameSite: str = ""

class ExportRequest(BaseModel):
    rows: list[dict]
    format: str = "xlsx"
    headers: list[str] | None = None

class RegisterElementItem(BaseModel):
    dedupKey: str = ""
    outerHTML: str = ""
    selector: str = ""
    xpath: str = ""
    source: str = ""
    tag: str = ""
    text: str = ""
    className: str = ""
    elementId: str = ""
    href: str = ""
    src: str = ""
    page_url: str = ""
    clean_selector: str = ""
    snapshot_id: int = 0

class RegisterRequest(BaseModel):
    elements: list[RegisterElementItem]

class ElementChainRequest(BaseModel):
    element_ids: list[str]
    chain_type: str = "css"
    deepest_selector: str = ""
    fields: list[dict] = []
    child_delim: str = ""         # 子节点分隔符

# ──────── Health ────────
@app.get("/api/health")
async def health():
    return {"status": "ok"}

# ──────── 前端日志 → 黑窗 ────────
class LogRequest(BaseModel):
    msg: str = ""

@app.post("/api/log")
async def frontend_log(req: LogRequest):
    logger.info(f"[前端] {req.msg}")
    return {"ok": True}

# ──────── 解析 API ────────
@app.post("/api/parse/all")
async def parse_all(req: ParseRequest):
    """一次性返回所有解析结果"""
    if not req.html or not req.html.strip():
        raise HTTPException(400, "HTML 为空")

    logger.info(f"解析 HTML, 长度={len(req.html)}")

    stats = get_source_stats(req.html)
    formatted = format_html(req.html)
    dom = build_dom_tree(req.html, max_depth=req.max_depth, max_text_len=req.max_text_len, max_children=getattr(req, 'max_children', 200))
    scripts = extract_scripts(req.html)

    return {
        "stats": stats,
        "formatted_html": formatted,
        "dom_tree": dom,
        "scripts": scripts,
    }

@app.post("/api/parse/html")
async def parse_html(req: ParseRequest):
    if not req.html or not req.html.strip():
        raise HTTPException(400, "HTML 为空")
    return {
        "stats": get_source_stats(req.html),
        "formatted_html": format_html(req.html),
    }

@app.post("/api/parse/dom")
async def parse_dom(req: ParseRequest):
    if not req.html or not req.html.strip():
        raise HTTPException(400, "HTML 为空")
    return {"dom_tree": build_dom_tree(req.html, max_depth=req.max_depth, max_text_len=req.max_text_len, max_children=getattr(req, 'max_children', 200))}

@app.post("/api/parse/scripts")
async def parse_scripts(req: ParseRequest):
    if not req.html or not req.html.strip():
        raise HTTPException(400, "HTML 为空")
    scripts = extract_scripts(req.html)
    # 尝试解析内嵌JSON
    for s in scripts:
        if s.get("content") and not s.get("src"):
            json_data = parse_json_from_text(s["content"])
            s["parsed_json"] = json_data
    return {"scripts": scripts}

# ──────── 提取 API ────────
@app.post("/api/extract/xpath")
async def extract_xpath(req: ParseRequest):
    """XPath 提取"""
    if not req.html or not req.html.strip():
        raise HTTPException(400, "HTML 为空")
    if not req.query:
        raise HTTPException(400, "XPath 查询表达式为空")
    return xpath_query(req.html, req.query, req.child_delim,
                       max_text_len=req.max_text_len, max_results=req.max_results,
                       expand_children=req.expand_children)

@app.post("/api/extract/css")
async def extract_css(req: ParseRequest):
    """CSS 选择器提取"""
    if not req.html or not req.html.strip():
        raise HTTPException(400, "HTML 为空")
    if not req.query:
        raise HTTPException(400, "CSS 选择器为空")
    result = css_query(req.html, req.query, req.child_delim,
                       max_text_len=req.max_text_len, max_results=req.max_results,
                       expand_children=req.expand_children)
    logger.info(f"[CSS] '{req.query[:80]}' → {result.get('count',0)} 条")
    return result

@app.post("/api/extract/regex")
async def extract_regex(req: ParseRequest):
    """正则提取"""
    if not req.html or not req.html.strip():
        raise HTTPException(400, "HTML 为空")
    if not req.query:
        raise HTTPException(400, "正则表达式为空")
    return regex_search(req.html, req.query)

@app.post("/api/extract/jsonpath")
async def extract_jsonpath(req: ParseRequest):
    """JSONPath 提取（在HTML中搜索JSON数据块，然后用JSONPath查询）"""
    if not req.html or not req.html.strip():
        raise HTTPException(400, "HTML 为空")
    if not req.query:
        raise HTTPException(400, "JSONPath 查询表达式为空")
    return jsonpath_query(req.html, req.query)

@app.post("/api/extract/chain")
async def extract_chain(req: ChainRequest):
    """链路提取（走 Python 后端 lxml 解析，解决虚拟列表 DOM 不全问题）"""
    logger.info(f"[链路] 收到请求: selector={req.deepest_selector[:120]} fields={len(req.fields)} html_len={len(req.html)}")
    if not req.html or not req.html.strip():
        raise HTTPException(400, "HTML 为空")
    if not req.deepest_selector:
        raise HTTPException(400, "deepest_selector 为空")
    if not req.fields:
        raise HTTPException(400, "fields 为空")
    result = chain_extract(req.html, req.chain_type, req.deepest_selector, req.fields, req.child_delim)
    logger.info(f"[链路] 结果: {result.get('totalRows',0)} 行, targets={result.get('_debug',{}).get('target_count','?')}")
    return result

@app.post("/api/extract/trace")
async def extract_trace(req: ChainRequest):
    """从 HTML 中用 lxml 做溯源，返回祖先链（不依赖浏览器 DOM）"""
    if not req.html or not req.html.strip():
        raise HTTPException(400, "HTML 为空")
    if not req.deepest_selector:
        raise HTTPException(400, "deepest_selector 为空")
    return trace_chain_backend(req.html, req.chain_type or 'css', req.deepest_selector)

# ──────── Cookie 管理 ────────
@app.get("/api/cookies/{domain}")
async def get_cookies(domain: str):
    """获取已保存的 cookie"""
    safe_domain = _safe_path_component(domain)
    file = COOKIES_DIR / f"{safe_domain}.json"
    if not file.exists():
        return {"cookies": [], "message": "未找到已保存的Cookie"}
    try:
        data = json.loads(file.read_text(encoding="utf-8"))
        # 检查过期
        import time
        now = time.time()
        valid = []
        expired = []
        for c in data.get("cookies", []):
            exp = c.get("expirationDate", 0)
            if exp == 0 or exp > now:
                valid.append(c)
            else:
                expired.append(c["name"])
        if expired:
            logger.info(f"Cookie 已过期: {expired}")
            file.write_text(json.dumps({"cookies": valid}, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"cookies": valid, "expired": expired}
    except Exception as e:
        return {"cookies": [], "error": str(e)}

@app.post("/api/cookies/{domain}")
async def save_cookies(domain: str, cookies: list[CookieItem]):
    """保存 cookie"""
    safe_domain = _safe_path_component(domain)
    file = COOKIES_DIR / f"{safe_domain}.json"
    data = {"cookies": [c.model_dump() for c in cookies], "saved_at": __import__("time").strftime("%Y-%m-%d %H:%M:%S")}
    file.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    logger.info(f"已保存 {len(cookies)} 条 Cookie → {domain}")
    return {"ok": True, "count": len(cookies), "domain": domain}

# ──────── 导出 ────────
@app.post("/api/export/excel")
async def export_excel(req: ExportRequest):
    """导出数据（支持 xlsx/csv/json/html）"""
    try:
        import io
        import base64

        if not req.rows:
            raise HTTPException(400, "没有可导出的数据")

        # 清洗所有行中的不可见字符和异常空白
        rows = _normalize_rows(req.rows)

        # 确定列头和顺序（过滤 _ 前缀的内部字段）
        if req.headers and len(req.headers) > 0:
            all_keys = [k for k in req.headers if not k.startswith('_')]
        else:
            all_keys = [k for k in rows[0].keys() if not k.startswith('_')]

        fmt = req.format.lower()

        if fmt == "xlsx":
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "解析结果"

            header_fill = PatternFill(start_color="7C5CFC", end_color="7C5CFC", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )

            for ci, k in enumerate(all_keys, 1):
                cell = ws.cell(row=1, column=ci, value=k)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            for ri, row in enumerate(rows, 2):
                for ci, k in enumerate(all_keys, 1):
                    val = row.get(k, "")
                    if isinstance(val, (dict, list)):
                        val = json.dumps(val, ensure_ascii=False)
                    val = normalize_text(str(val))
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")

            for ci, k in enumerate(all_keys, 1):
                max_len = len(k)
                for ri in range(2, len(rows) + 2):
                    val = str(ws.cell(row=ri, column=ci).value or "")
                    max_len = max(max_len, len(val))
                ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max_len + 6, 60)

            ws.freeze_panes = "A2"

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            b64 = base64.b64encode(buf.read()).decode()
            return {"ok": True, "filename": "export.xlsx", "data": b64, "format": "xlsx"}

        elif fmt == "csv":
            import csv as csv_mod
            buf = io.StringIO()
            writer = csv_mod.writer(buf)
            writer.writerow(all_keys)
            for row in rows:
                writer.writerow([str(row.get(k, "")) for k in all_keys])
            result = buf.getvalue()
            # 前端期望 base64
            b64 = base64.b64encode(result.encode("utf-8-sig")).decode()
            return {"ok": True, "filename": "export.csv", "data": b64, "format": "csv"}

        elif fmt == "json":
            result = json.dumps(rows, ensure_ascii=False, indent=2)
            b64 = base64.b64encode(result.encode("utf-8")).decode()
            return {"ok": True, "filename": "export.json", "data": b64, "format": "json"}

        elif fmt == "html":
            html = '<!DOCTYPE html>\n<html lang="zh-CN">\n<head>\n<meta charset="UTF-8">\n'
            html += '<title>导出数据</title>\n'
            html += '<style>\n'
            html += 'body{font-family:-apple-system,BlinkMacSystemFont,"Microsoft YaHei",sans-serif;margin:20px;color:#333}\n'
            html += 'table{border-collapse:collapse;width:100%}\n'
            html += 'th{background:#7c5cfc;color:#fff;padding:10px 12px;text-align:left;border:1px solid #6a4de6}\n'
            html += 'td{padding:8px 12px;border:1px solid #ddd}\n'
            html += 'tr:nth-child(even){background:#f7f7f7}\n'
            html += '</style>\n</head>\n<body>\n<table>\n<thead>\n<tr>\n'
            for k in all_keys:
                html += f'<th>{escape_html(k)}</th>\n'
            html += '</tr>\n</thead>\n<tbody>\n'
            for row in rows:
                html += '<tr>\n'
                for k in all_keys:
                    html += f'<td>{escape_html(str(row.get(k, "")))}</td>\n'
                html += '</tr>\n'
            html += '</tbody>\n</table>\n</body>\n</html>'
            b64 = base64.b64encode(html.encode("utf-8")).decode()
            return {"ok": True, "filename": "export.html", "data": b64, "format": "html"}

        elif fmt == "md":
            lines = []
            lines.append("| " + " | ".join(str(k) for k in all_keys) + " |")
            lines.append("|" + "|".join("---" for _ in all_keys) + "|")
            for row in rows:
                vals = []
                for k in all_keys:
                    v = str(row.get(k, ""))
                    v = v.replace("\\", "\\\\").replace("|", "\\|").replace("\n", "<br>")
                    vals.append(v)
                lines.append("| " + " | ".join(vals) + " |")
            result = "\n".join(lines)
            b64 = base64.b64encode(result.encode("utf-8")).decode()
            return {"ok": True, "filename": "export.md", "data": b64, "format": "md"}

        else:
            raise HTTPException(400, f"不支持的格式: {fmt}")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"导出失败: {e}")
        raise HTTPException(500, f"导出失败: {e}")

# 预编译 Unicode 清洗正则（导出时使用）
_ZSP = re.compile(r"[   -   　  ]")
_CF = re.compile(r"[\u00AD\u0600-\u0605\u061C\u06DD\u070F\u08E2\u180E\u200B-\u200F\u202A-\u202E\u2060-\u2064\u2066-\u206F\uFEFF\uFFF9-\uFFFB\U000110BD\U000110CD\U00013430-\U0001343F\U0001BCA0-\U0001BCA3\U0001D173-\U0001D17A\U000E0001\U000E0020-\U000E007F]")
_CC = re.compile(r"[\u0000-\u0008\u000E-\u001F\u007F-\u009F]")
_PUA = re.compile(r"[\uE000-\uF8FF\U000F0000-\U000FFFFD\U00100000-\U0010FFFD\uFFFE\uFFFF\U0001FFFE\U0001FFFF\U0002FFFE\U0002FFFF\U0003FFFE\U0003FFFF\U0004FFFE\U0004FFFF\U0005FFFE\U0005FFFF\U0006FFFE\U0006FFFF\U0007FFFE\U0007FFFF\U0008FFFE\U0008FFFF\U0009FFFE\U0009FFFF\U000AFFFE\U000AFFFF\U000BFFFE\U000BFFFF\U000CFFFE\U000CFFFF\U000DFFFE\U000DFFFF\U000EFFFE\U000EFFFF\U000FFFFE\U000FFFFF\U0010FFFE\U0010FFFF]")
def normalize_text(s: str) -> str:
    """清理不可见字符和异常空白，避免导出文件中出现"口"等乱码"""
    if not s:
        return ""
    s = str(s)
    # 1) Zs/Zl/Zp 空白分隔符（含 ideographic space U+3000）→ 普通空格
    s = _ZSP.sub(" ", s)
    # 2) 移除不可见格式化字符（Cf 类别：零宽字符、双向控制、软连字符、BOM 等）
    s = _CF.sub("", s)
    # 3) 移除控制字符（Cc），保留基本空白（后续 \s+ 统一压缩）
    s = _CC.sub("", s)
    # 4) 移除私有区（PUA）和无标准字形的特殊字符
    s = _PUA.sub("", s)
    # 5) 合并连续空白，去除首尾
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _normalize_rows(rows: list[dict]) -> list[dict]:
    """递归清洗所有字符串值中的不可见字符，返回新列表"""
    result = []
    for row in rows:
        clean = {}
        for k, v in row.items():
            if isinstance(v, str):
                v = normalize_text(v)
            elif isinstance(v, (dict, list)):
                v = normalize_text(json.dumps(v, ensure_ascii=False))
            clean[k] = v
        result.append(clean)
    return result

def escape_html(text: str) -> str:
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")

# ──────── 保存源码到本地 ────────
@app.post("/api/save/source")
async def save_source(req: ParseRequest):
    """保存HTML源码到本地文件"""
    import tempfile
    try:
        tmp = Path(tempfile.gettempdir()) / "webparser_source.html"
        tmp.write_text(req.html, encoding="utf-8")
        return {"ok": True, "path": str(tmp), "size": len(req.html)}
    except Exception as e:
        raise HTTPException(500, f"保存失败: {e}")

# ──────── 已注册元素 ────────

@app.post("/api/elements/register")
async def register_elements(req: RegisterRequest):
    """注册元素到后端（SQLite 持久化），按 dedupKey 去重"""
    payload = []
    for elem in req.elements:
        payload.append({
            "dedupKey": elem.dedupKey,
            "outerHTML": elem.outerHTML,
            "selector": elem.selector,
            "xpath": elem.xpath,
            "source": elem.source,
            "tag": elem.tag,
            "text": elem.text,
            "className": elem.className,
            "elementId": elem.elementId,
            "href": elem.href,
            "src": elem.src,
            "page_url": elem.page_url,
            "clean_selector": elem.clean_selector,
            "snapshot_id": elem.snapshot_id,
        })
    return db.register_elements(payload)

@app.get("/api/elements")
async def list_elements():
    """列出所有已注册元素（从 SQLite）"""
    elements = db.list_elements()
    return {"ok": True, "elements": elements, "total": len(elements)}

@app.delete("/api/elements")
async def clear_elements(element_ids: list[str] | None = None):
    """清空注册表，或删除指定元素"""
    return db.clear_elements(element_ids)

# ═══════════════════════════════════════════
#  元素链路提取
# ═══════════════════════════════════════════


@app.post("/api/elements/batch")
async def upsert_element_batch(data: dict):
    """批量存储注册元素行数组 {page_url, snapshot_id, headers, rows}"""
    page_url = data.get("page_url", "")
    snapshot_id = data.get("snapshot_id", 0)
    if not page_url:
        raise HTTPException(400, "page_url required")
    return db.upsert_element_batch(page_url, snapshot_id, {
        "headers": data.get("headers", []),
        "rows": data.get("rows", []),
    })

@app.get("/api/elements/batch")
async def get_element_batch(url: str = "", snapshot_id: int = 0):
    """取批量注册的行数组（优先 snapshot_id，回退 url 兼容旧数据）"""
    if snapshot_id:
        result = db.get_element_batch(snapshot_id)
        if result is not None:
            return {"ok": True, "data": result}
    if url:
        result = db.get_element_batch(0)  # fallback
        return {"ok": True, "data": result} if result else {"ok": True, "data": None}
    raise HTTPException(400, "url or snapshot_id required")
@app.post("/api/elements/chain")
async def chain_from_elements(req: ElementChainRequest):
    """对已注册元素执行链路提取"""
    results_rows = []
    all_headers = []
    total = 0
    all_elements = db.list_elements()
    registry = {e["id"]: e for e in all_elements}

    for eid in req.element_ids:
        entry = registry.get(eid)
        if not entry:
            continue

        html_str = entry.get("html", "")
        if not html_str:
            continue

        # 如果提供了 deepest_selector，使用原始链提取逻辑
        if req.deepest_selector:
            result = chain_extract(
                raw_html=html_str,
                chain_type=req.chain_type,
                deepest_selector=req.deepest_selector,
                fields=req.fields,
                child_delim=req.child_delim,
            )
        else:
            # 没有 deepest_selector → 元素本身是目标
            result = chain_extract(
                raw_html=html_str,
                chain_type=req.chain_type,
                deepest_selector=req.deepest_selector or "*",
                fields=req.fields,
                child_delim=req.child_delim,
            )

        for row in result.get("rows", []):
            row["_elem_id"] = eid
            results_rows.append(row)
        total += result.get("totalRows", 0)
        if not all_headers and result.get("headers"):
            all_headers = result["headers"]

    return {
        "ok": True,
        "rows": results_rows,
        "headers": all_headers,
        "totalRows": total,
    }

# ═══════════════════════════════════════════
#  页面快照（翻页批量提取）
# ═══════════════════════════════════════════

class PageSnapshotSaveRequest(BaseModel):
    url: str = ""
    html: str

@app.post("/api/page-snapshots/save")
async def save_page_snapshot(req: PageSnapshotSaveRequest):
    """注册元素时顺手保存当前页 HTML"""
    logger.info(f"[快照] 收到保存请求: url={req.url[:80]} html_len={len(req.html)}")
    if not req.html or not req.html.strip():
        logger.warning("[快照] HTML为空，拒绝保存")
        raise HTTPException(400, "HTML 为空")
    result = db.save_page_snapshot(req.url, req.html)
    logger.info(f"[快照] 已保存: total={result['total_snapshots']}")
    return result

@app.get("/api/page-snapshots/list")
async def list_page_snapshots():
    """列出所有页面快照（不含 HTML）"""
    snapshots = db.list_page_snapshots()
    return {"ok": True, "snapshots": snapshots, "total": len(snapshots)}

@app.get("/api/page-snapshots/{snapshot_id}/html")
async def get_snapshot_html(snapshot_id: int):
    """获取单个快照的 HTML"""
    html = db.get_page_snapshot_html(snapshot_id)
    if html is None:
        raise HTTPException(404, "快照不存在")
    return {"ok": True, "html": html, "id": snapshot_id}

@app.delete("/api/page-snapshots")
async def clear_page_snapshots():
    """清空所有页面快照（开始新一批翻页时调用）"""
    logger.info("[快照] 收到清空请求")
    return db.clear_page_snapshots()

# ═══════════════════════════════════════════
#  链路方案数据持久化
# ═══════════════════════════════════════════

class ChainDataSaveRequest(BaseModel):
    scheme_name: str
    rows: list[dict]
    headers: list[str]


# ── 方案存储 ──

@app.post("/api/schemes")
async def save_scheme(data: dict):
    """保存/更新方案"""
    name = data.get("name", "")
    schema = data.get("schema", {})
    if not name:
        raise HTTPException(400, "name required")
    return db.save_scheme(name, schema)

@app.get("/api/schemes/{name}")
async def load_scheme(name: str):
    """加载方案"""
    result = db.load_scheme(name)
    if result is None:
        raise HTTPException(404, "scheme not found")
    return result

@app.delete("/api/schemes/{name}")
async def delete_scheme(name: str):
    """删除方案"""
    return db.delete_scheme(name)

@app.get("/api/schemes")
async def list_schemes():
    """列出所有方案名"""
    return {"names": db.list_schemes()}

@app.post("/api/chain-data/save")
async def save_chain_data(req: ChainDataSaveRequest):
    logger.info(f"[保存] scheme={req.scheme_name} headers={req.headers} rows={len(req.rows)}")
    if req.rows and len(req.rows) > 0:
        logger.info(f"[保存] 第1行: { {k: str(v)[:40] for k, v in list(req.rows[0].items())[:8]} }")
    return db.save_chain_data(req.scheme_name, req.rows, req.headers)

@app.post("/api/debug")
async def debug_log(req: dict = Body(None)):
    """前端调试日志 → 后端黑窗"""
    msg = (req or {}).get("msg", "")
    logger.info(f"[前端] {msg}")
    return {"ok": True}

@app.get("/api/chain-data/query")
async def query_chain_data(schemes: str = "", link_col: str = "", link_cols: str = "", mode: str = ""):
    names = [n.strip() for n in schemes.split(",") if n.strip()]
    if not names:
        return {"rows": [], "headers": [], "totalRows": 0}
    if mode == "vertical" and len(names) >= 2:
        return db.merge_schemes_vertical(names)
    per_scheme = [c.strip() for c in link_cols.split(",") if c.strip()] if link_cols else []
    return db.get_chain_data(names, link_col=link_col, link_cols=per_scheme)

@app.delete("/api/chain-data/{scheme_name}")
async def delete_chain_data(scheme_name: str):
    return db.delete_chain_data(scheme_name)

# ── 共存合并 ──

@app.post("/api/merge/inline")
async def merge_inline(data: dict):
    """预览用：传当前链提取结果+批量数据，直接合并"""
    return db.merge_rows(
        data.get("chain_rows", []),
        data.get("chain_headers", []),
        data.get("batch_rows", []),
        data.get("batch_headers", []),
    )

@app.post("/api/merge/query")
async def merge_query(data: dict):
    """保存并查询用：从库读取链数据+批量数据，合并"""
    return db.merge_chain_and_batch(
        data.get("scheme_name", ""),
        data.get("snapshot_id", 0),
    )

@app.put("/api/chain-data/update-row")
async def update_chain_row(scheme_name: str, row_index: int, data: dict):
    return db.update_chain_row(scheme_name, row_index, data)

@app.get("/api/chain-data/list")
async def list_chain_data():
    return {"schemes": db.list_chain_schemes_with_data()}

# ═══════════════════════════════════════════
#  采集数据管线 (SQLite 持久化)
# ═══════════════════════════════════════════

class CollectIngestRequest(BaseModel):
    source: str = ""          # "scroll" | "api"
    url: str = ""
    rows: list[dict] = []
    collect_id: str = ""      # 为空则自动生成

class CollectCleanRequest(BaseModel):
    collect_id: str
    rules: list[dict] = []    # [{type: "trim"|"strip_empty"|"regex"|"dedup", field, pattern, targetType}]

@app.post("/api/collect/ingest")
async def collect_ingest(req: CollectIngestRequest):
    """接收采集数据（SQLite 持久化）"""
    if not req.rows:
        return {"ok": True, "total": 0, "collect_id": ""}

    # 自动生成 collect_id
    cid = req.collect_id
    if not cid:
        import hashlib
        raw = json.dumps(req.rows, ensure_ascii=False, sort_keys=True)
        cid = hashlib.md5(raw.encode()).hexdigest()[:12]

    return db.ingest_collected(cid, req.source, req.url, req.rows)

@app.post("/api/collect/clean")
async def collect_clean(req: CollectCleanRequest):
    """清洗已采集数据"""
    return db.clean_collected(req.collect_id, req.rules)

@app.get("/api/collect/data")
async def collect_list(collect_id: str = ""):
    """列出已持久化的采集数据"""
    if collect_id:
        rows = db.load_collected(collect_id)
        meta = db.load_collected_meta(collect_id)
        return {"ok": True, "collect_id": collect_id, "rows": rows,
                "total": len(rows), "source": meta.get("source", "")}
    collections = db.list_collections()
    return {"ok": True, "collections": collections}

@app.delete("/api/collect/data")
async def collect_delete(collect_id: str = ""):
    """删除采集数据"""
    return db.delete_collected(collect_id)

# ═══════════════════════════════════════════
#  导出方案管理
# ═══════════════════════════════════════════

class SchemaSaveRequest(BaseModel):
    name: str
    fields: list[dict]

@app.post("/api/schemas")
async def save_schema(req: SchemaSaveRequest):
    """保存导出方案"""
    return db.save_schema(req.name, req.fields)

@app.get("/api/schemas")
async def list_schemas():
    """列出所有方案"""
    return {"ok": True, "schemas": db.list_schemas()}

@app.delete("/api/schemas/{name}")
async def delete_schema(name: str):
    """删除方案"""
    return db.delete_schema(name)

# ═══════════════════════════════════════════
#  应用设置
# ═══════════════════════════════════════════

@app.get("/api/settings")
async def get_settings():
    """获取所有设置"""
    return {"ok": True, "settings": db.load_all_settings()}

@app.post("/api/settings")
async def save_settings(settings: dict):
    """批量保存设置"""
    return db.save_all_settings(settings)

@app.get("/api/settings/{key}")
async def get_setting(key: str):
    """获取单个设置"""
    value = db.load_setting(key)
    return {"ok": True, "key": key, "value": value}

@app.put("/api/settings/{key}")
async def set_setting(key: str, value: dict | str | int | float | bool | list = None):
    """保存单个设置"""
    return db.save_setting(key, value)

# ════════════════════════════════════════════════
# 跨平台比价 API
# ════════════════════════════════════════════════

from fastapi import UploadFile, File, Form
from typing import Optional

class IngestRequest(BaseModel):
    platform: str
    url: str
    title: str = ""
    price: float = 0
    shop_name: str = ""
    main_image_url: str = ""
    attrs: dict = None
    stock_info: str = ""
    shipping: str = ""
    description: str = ""
    image_urls: list = None

@app.post("/api/price-compare/ingest")
async def ingest_product(req: IngestRequest):
    """入库一个商品（从链接）"""
    pid = product_pipeline.ingest_product(
        platform=req.platform,
        url=req.url,
        title=req.title,
        price=req.price,
        shop_name=req.shop_name,
        main_image_url=req.main_image_url,
        attrs=req.attrs,
        stock_info=req.stock_info,
        shipping=req.shipping,
        description=req.description,
        image_urls=req.image_urls,
    )
    return {"ok": True, "product_id": pid}

@app.post("/api/price-compare/ingest-batch")
async def ingest_products(products: list[dict]):
    """批量入库"""
    ids = product_pipeline.ingest_products(products)
    return {"ok": True, "product_ids": ids, "count": len(ids)}

@app.post("/api/price-compare/search-by-image")
async def search_by_image(file: UploadFile = File(...), top_k: int = 20, platform: str = "", skip_rembg: str = "0"):
    """以图搜图（可选按平台过滤：taobao/jd/pdd/1688/tmall）"""
    data = await file.read()
    skip_rembg_flag = skip_rembg == "1"
    try:
        results = product_pipeline.search_by_image(
            image_data=data, top_k=top_k,
            platform=platform if platform else None,
            skip_rembg=skip_rembg_flag
        )
        return {"ok": True, "results": results, "count": len(results)}
    except Exception as e:
        import traceback
        logger.error(f"[search-by-image] {type(e).__name__}: {e}")
        logger.error(traceback.format_exc())
        return {"ok": False, "error": f"{type(e).__name__}: {e}", "results": [], "count": 0}

@app.get("/api/price-compare/search-by-url")
async def search_by_url(url: str, top_k: int = 20):
    """根据商品链接搜同款"""
    try:
        results = product_pipeline.search_by_url(url, top_k=top_k)
        return {"ok": True, "results": results, "count": len(results)}
    except Exception as e:
        import traceback
        logger.error(f"[search-by-url] {type(e).__name__}: {e}")
        logger.error(traceback.format_exc())
        return {"ok": False, "error": f"{type(e).__name__}: {e}", "results": [], "count": 0}

@app.post("/api/price-compare/search-by-text")
async def search_by_text(text: str = Form(...), top_k: int = Form(20), 
                          min_score: float = Form(0.35), platform: str = Form("")):
    """中文文字搜图（利用 Chinese-CLIP 文本编码器）
    min_score 默认 0.35（跨模态匹配天然低于图搜图）"""
    try:
        results = product_pipeline.search_by_text(
            text=text, top_k=top_k, min_score=min_score,
            platform=platform if platform else None
        )
        return {"ok": True, "results": results, "count": len(results)}
    except Exception as e:
        import traceback
        logger.error(f"[search-by-text] {type(e).__name__}: {e}")
        logger.error(traceback.format_exc())
        return {"ok": False, "error": f"{type(e).__name__}: {e}", "results": [], "count": 0}

@app.get("/api/price-compare/stats")
async def get_price_compare_stats():
    """获取比价库状态"""
    stats = product_pipeline.get_stats()
    return {"ok": True, **stats}

@app.post("/api/price-compare/rebuild-vectors")
async def rebuild_all_vectors():
    """重新为所有已入库商品生成向量（后台线程 + 进度轮询）"""
    import sqlite3 as _sqlite3
    import uuid as _uuid, threading as _threading, time as _time
    import traceback as _traceback

    job_id = _uuid.uuid4().hex[:8]
    _import_jobs[job_id] = {
        "status": "running",
        "progress": {"done": 0, "total": 0, "current": "准备中...", "elapsed_s": 0, "eta_s": 0, "step": "init"},
        "result": None,
    }

    def _run():
        job = _import_jobs[job_id]
        try:
            conn2 = _sqlite3.connect(str(db.DB_PATH))
            conn2.row_factory = _sqlite3.Row
            rows = conn2.execute(
                "SELECT id, title, local_image, platform FROM products WHERE status='active'"
            ).fetchall()
            conn2.close()

            total = len(rows)
            job["progress"]["total"] = total
            start = _time.time()
            success = skipped = fail = 0

            for i, r in enumerate(rows):
                pid = r["id"]

                elapsed = _time.time() - start
                avg = elapsed / (i + 1) if i > 0 else 0
                eta = avg * (total - (i + 1)) if avg > 0 else 0
                job["progress"] = {
                    "done": i, "total": total,
                    "current": r["title"][:40] if r["title"] else f"#{pid}",
                    "elapsed_s": round(elapsed, 1),
                    "eta_s": round(eta, 1),
                    "step": "重建中",
                }

                try:
                    result = product_pipeline.rebuild_product_vectors(pid)
                    if result.get("ok"):
                        success += 1
                    else:
                        fail += 1
                except Exception as e:
                    logger.error(f"[rebuild] #{pid} 向量重建失败: {e}")
                    fail += 1

            total_elapsed = _time.time() - start
            job["progress"]["done"] = total
            job["progress"]["step"] = "完成"
            job["progress"]["elapsed_s"] = round(total_elapsed, 1)
            job["progress"]["eta_s"] = 0
            job["status"] = "done"
            job["result"] = {
                "total": total, "success": success, "skipped": skipped, "failed": fail,
                "vector_count": product_pipeline.vector_store.get_count(),
            }
            logger.info(f"[rebuild] Done: {success} ok, {skipped} skip, {fail} fail in {total_elapsed:.1f}s")
        except Exception as e:
            job["status"] = "error"
            job["result"] = {"error": f"{type(e).__name__}: {e}", "traceback": _traceback.format_exc()}

    _threading.Thread(target=_run, daemon=True).start()
    return {"ok": True, "job_id": job_id}

@app.get("/api/price-compare/products")
async def list_products(platform: str = None, limit: int = 100):
    """列出已入库商品"""
    products = db.get_all_products(platform=platform, limit=limit)
    return {"ok": True, "products": products, "count": len(products)}

@app.get("/api/price-compare/products/search")
async def search_products(q: str = "", platform: str = None,
                          limit: int = 50, page: int = 1):
    """搜索商品（关键词 + 平台筛选 + 分页）"""
    offset = (page - 1) * limit
    products = db.search_products(query=q, platform=platform, limit=limit, offset=offset)
    total = db.count_search_results(query=q, platform=platform)
    total_pages = max(1, (total + limit - 1) // limit) if total > 0 else 1
    return {
        "ok": True,
        "products": products,
        "total": total,
        "page": page,
        "total_pages": total_pages,
        "limit": limit,
    }

@app.get("/api/price-compare/products/{product_id}")
async def get_product(product_id: int):
    """单条商品详情"""
    p = db.get_product_by_id(product_id)
    if not p:
        raise HTTPException(status_code=404, detail="商品不存在")
    return {"ok": True, "product": p}

class ProductUpdateBody(BaseModel):
    title: str = None
    price: float = None
    original_price: float = None
    shop_name: str = None
    main_image_url: str = None
    platform: str = None
    platform_url: str = None
    attrs: dict = None
    stock_info: str = None
    shipping: str = None
    description: str = None
    image_urls: list = None
    skus: list = None

@app.put("/api/price-compare/products/{product_id}")
async def update_product(product_id: int, body: ProductUpdateBody):
    """更新商品字段（传哪些字段就更新哪些）。若涉及图片变更，后台自动重建向量。"""
    p = db.get_product_by_id(product_id)
    if not p:
        raise HTTPException(status_code=404, detail="商品不存在")
    fields = {k: v for k, v in body.model_dump().items() if v is not None}
    if not fields:
        return {"ok": False, "error": "没有提供要更新的字段"}
    db.update_product(product_id, **fields)

    # 若修改了图片相关字段（main_image_url 或 image_urls），后台重建向量
    if "main_image_url" in fields or "image_urls" in fields:
        threading.Thread(
            target=product_pipeline.rebuild_product_vectors,
            args=(product_id,),
            daemon=True,
        ).start()

    # 重新返回最新数据
    updated = db.get_product_by_id(product_id)
    return {"ok": True, "product": updated}

@app.delete("/api/price-compare/products/{product_id}")
async def delete_product(product_id: int):
    """删除商品（软删除，保留数据库记录但标记为 deleted）"""
    ok = db.delete_product(product_id)
    if not ok:
        raise HTTPException(status_code=404, detail="商品不存在或已删除")
    return {"ok": True}

class BatchDeleteBody(BaseModel):
    product_ids: list[int]

@app.post("/api/price-compare/products/batch-delete")
async def batch_delete_products(body: BatchDeleteBody):
    """批量软删除"""
    if not body.product_ids:
        return {"ok": False, "error": "product_ids 不能为空"}
    deleted = db.batch_delete_products(body.product_ids)
    return {"ok": True, "deleted": deleted}

@app.post("/api/price-compare/products/repair-platforms")
async def repair_platforms(data: dict = None):
    """重新检测商品平台（根据 platform_url 自动识别）。
    不传 body → 修复所有 platform 为空的记录；传 product_ids → 只修复指定 ID。"""
    try:
        product_ids = data.get("product_ids", []) if data else []
        if product_ids:
            products = db.get_products_by_ids(product_ids)
        else:
            # 修复所有 platform 为空的活跃商品
            import sqlite3
            conn = sqlite3.connect(str(db.DB_PATH))
            conn.row_factory = sqlite3.Row
            rows = conn.execute(
                "SELECT * FROM products WHERE status='active' AND (platform='' OR platform='unknown')"
            ).fetchall()
            conn.close()
            products = [db._product_row_to_dict(r) for r in rows]

        repaired = 0
        skipped = 0
        for p in products:
            detected = detect_platform_from_url(p.get("platform_url", ""))
            if detected and detected != p.get("platform", ""):
                try:
                    db.update_product(p["id"], platform=detected)
                    repaired += 1
                except Exception:
                    skipped += 1

        return {"ok": True, "total": len(products), "repaired": repaired, "skipped": skipped}
    except Exception as e:
        return {"ok": False, "error": str(e)}

class IngestFromChainBody(BaseModel):
    scheme_name: str

@app.post("/api/price-compare/ingest-from-chain")
async def ingest_from_chain(body: IngestFromChainBody):
    """从链路提取结果自动入库到比价库"""
    return product_pipeline.ingest_from_chain_data(body.scheme_name)

# ═══════════════════════════════════════════════
# 跨页面数据桥：比价控制台 ⇄ 网页解析器 query
# ═══════════════════════════════════════════════
_injected_compare_data = []

class CompareInjectBody(BaseModel):
    results: list

@app.post("/api/query/inject-price-compare")
async def inject_price_compare(body: CompareInjectBody):
    """比价控制台把搜索结果注入，供网页解析器 query 面板拉取"""
    global _injected_compare_data
    _injected_compare_data = body.results
    return {"ok": True, "count": len(_injected_compare_data)}

@app.get("/api/query/pull-price-compare")
async def pull_price_compare():
    """网页解析器拉取比价数据注入 query 表格"""
    return {"ok": True, "results": _injected_compare_data, "count": len(_injected_compare_data)}

@app.post("/api/price-compare/rebuild-vectors")
async def rebuild_vectors(data: dict = None):
    try:
        product_ids = (data or {}).get("product_ids", [])
        if not product_ids:
            all_products = db.get_products_by_ids(None) or []
            product_ids = [p.get("id") for p in all_products if p.get("status") == "active"]
        results = []
        for pid in product_ids:
            r = product_pipeline.rebuild_product_vectors(pid)
            results.append({"product_id": pid, "ok": r.get("ok", False), "count": r.get("count", 0)})
        return {"ok": True, "results": results}
    except Exception as e:
        import traceback
        return {"ok": False, "error": str(e), "traceback": traceback.format_exc()}
# ═══════════════════════════════════════════════
# 入库模板管理
# ═══════════════════════════════════════════════
TEMPLATES_DIR = os.path.join(os.path.dirname(__file__), "data", "templates")
os.makedirs(TEMPLATES_DIR, exist_ok=True)

@app.get("/api/price-compare/templates")
async def list_templates():
    """列出所有入库模板"""
    templates = []
    import glob as _glob
    for f in sorted(_glob.glob(os.path.join(TEMPLATES_DIR, "*.json"))):
        try:
            with open(f, encoding="utf-8") as fp:
                t = json.load(fp)
            templates.append(t)
        except Exception:
            pass
    return {"ok": True, "templates": templates}

class TemplateSave(BaseModel):
    name: str
    label: str
    platform: str = ""
    mappings: dict  # {"title": "标题列", "price": "价格列", ...}

@app.post("/api/price-compare/templates")
async def save_template(body: TemplateSave):
    """保存入库模板"""
    template = body.model_dump()
    template["name"] = template["name"].strip().lower().replace(" ", "_")
    path = os.path.join(TEMPLATES_DIR, f"{template['name']}.json")
    with open(path, "w", encoding="utf-8") as fp:
        json.dump(template, fp, ensure_ascii=False, indent=2)
    return {"ok": True, "template": template}

@app.delete("/api/price-compare/templates/{name}")
async def delete_template(name: str):
    """删除入库模板"""
    path = os.path.join(TEMPLATES_DIR, f"{name}.json")
    if os.path.exists(path):
        os.remove(path)
    return {"ok": True}

# ═══════════════════════════════════════════════
# 文件解析：Excel / CSV / JSON / SQLite → 行列表
# ═══════════════════════════════════════════════
def _parse_upload(file_bytes: bytes, filename: str) -> list[dict]:
    """统一解析上传文件，返回 [{"col1": val1, ...}, ...]"""
    ext = Path(filename).suffix.lower()

    if ext in (".csv", ".txt"):
        text = file_bytes.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [dict(row) for row in reader]

    if ext in (".xlsx", ".xls"):
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return []
        headers = [str(c or f"col_{i}") for i, c in enumerate(rows[0])]
        result = []
        for row in rows[1:]:
            d = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    d[headers[i]] = str(val) if val is not None else ""
            result.append(d)
        return result

    if ext == ".json":
        data = json.loads(file_bytes.decode("utf-8"))
        if isinstance(data, list):
            return data
        # 如果是 { "results": [...] } 嵌套结构，尝试提取
        if isinstance(data, dict):
            for v in data.values():
                if isinstance(v, list) and len(v) > 0 and isinstance(v[0], dict):
                    return v
            return []
        return []

    if ext in (".db", ".sqlite", ".sqlite3"):
        tmp = Path(os.path.join(TEMPLATES_DIR, "_temp_import.db"))
        tmp.write_bytes(file_bytes)
        try:
            conn = sqlite3.connect(str(tmp))
            cursor = conn.cursor()
            tables = cursor.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
            rows_list = []
            for (tbl,) in tables:
                cursor.execute(f"SELECT * FROM [{tbl}]")
                cols = [desc[0] for desc in cursor.description]
                for row in cursor.fetchall():
                    rows_list.append({cols[i]: str(row[i]) if row[i] is not None else "" for i in range(len(cols))})
            conn.close()
        finally:
            if tmp.exists():
                tmp.unlink()
        return rows_list

    raise HTTPException(400, f"不支持的文件格式: {ext}")

# ═══════════════════════════════════════════════
# 导入入库
# ═══════════════════════════════════════════════

SAMPLE_CSV_FIELDS = [
    "platform", "platform_url", "title", "price", "original_price",
    "shop_name", "main_image_url", "stock_info", "shipping",
    "description", "attrs_json", "image_urls", "skus",
]

SAMPLE_CSV_ROWS = [
    {
        "platform": "taobao",
        "platform_url": "https://item.taobao.com/item.htm?id=12345678901",
        "title": "夏季新款纯棉T恤 男女同款 宽松短袖",
        "price": "59.9",
        "original_price": "129",
        "shop_name": "潮流服饰旗舰店",
        "main_image_url": "https://img.alicdn.com/example/main.jpg",
        "stock_info": "库存 5000+",
        "shipping": "包邮",
        "description": "100%精梳棉，舒适透气，不起球不褪色",
        "attrs_json": '{"品牌":"潮流服饰","材质":"纯棉","尺码":"M,L,XL"}',
        "image_urls": '["https://img.alicdn.com/example/1.jpg","https://img.alicdn.com/example/2.jpg"]',
        "skus": '[{"color":"黑色","size":"M","price":59.9},{"color":"白色","size":"L","price":69.9}]',
    },
    {
        "platform": "jd",
        "platform_url": "https://item.jd.com/9876543210.html",
        "title": "无线蓝牙耳机 主动降噪 超长续航",
        "price": "199",
        "original_price": "399",
        "shop_name": "数码科技专营店",
        "main_image_url": "https://img.jd.com/example/main.jpg",
        "stock_info": "库存 1200+",
        "shipping": "京东物流·次日达",
        "description": "蓝牙5.3，ANC主动降噪，续航40小时",
        "attrs_json": '{"品牌":"数码科技","型号":"ANC-Pro","颜色":"黑色"}',
        "image_urls": '["https://img.jd.com/example/1.jpg","https://img.jd.com/example/2.jpg"]',
        "skus": '[{"color":"黑色","price":199},{"color":"白色","price":219}]',
    },
    {
        "platform": "pdd",
        "platform_url": "https://mobile.yangkeduo.com/goods.html?goods_id=555666777",
        "title": "厨房置物架 多层收纳架 微波炉架子",
        "price": "29.9",
        "original_price": "89",
        "shop_name": "居家好物小店",
        "main_image_url": "https://img.pdd.com/example/main.jpg",
        "stock_info": "库存 8000+",
        "shipping": "包邮·48小时发货",
        "description": "加厚碳钢材质，承重200斤，免打孔安装",
        "attrs_json": '{"材质":"碳钢","层数":"4层","颜色":"黑色"}',
        "image_urls": '["https://img.pdd.com/example/1.jpg"]',
        "skus": '[]',
    },
]


@app.get("/api/price-compare/sample-import-file")
async def download_sample_import_file():
    """返回批量入库示例 CSV 文件"""
    buf = io.StringIO()
    # UTF-8 BOM 确保 Excel 正确识别中文
    buf.write("\ufeff")
    writer = csv.writer(buf)
    writer.writerow(SAMPLE_CSV_FIELDS)
    for row in SAMPLE_CSV_ROWS:
        writer.writerow([row.get(f, "") for f in SAMPLE_CSV_FIELDS])

    csv_bytes = buf.getvalue().encode("utf-8")
    return Response(
        content=csv_bytes,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": "attachment; filename*=UTF-8''%E6%AF%94%E4%BB%B7%E6%89%B9%E9%87%8F%E5%85%A5%E5%BA%93_%E7%A4%BA%E4%BE%8B.csv"
        },
    )


@app.post("/api/price-compare/preview")
async def preview_import(file: UploadFile = File(...), template_name: str = Form("")):
    """预览：上传文件 + 指定模板 → 返回前 N 行映射结果"""
    file_bytes = await file.read()
    rows = _parse_upload(file_bytes, file.filename or "data.csv")

    template = None
    if template_name:
        tpl_path = os.path.join(TEMPLATES_DIR, f"{template_name}.json")
        if os.path.exists(tpl_path):
            with open(tpl_path, encoding="utf-8") as fp:
                template = json.load(fp)

    preview = []
    for i, row in enumerate(rows[:20]):
        mapped = {}
        if template:
            for target_field, source_col in template.get("mappings", {}).items():
                mapped[target_field] = str(row.get(source_col, "")).strip()
        else:
            mapped = {k: str(v).strip() for k, v in list(row.items())[:6]}
        preview.append(mapped)

    return {
        "ok": True,
        "total_rows": len(rows),
        "preview": preview,
        "columns": list(rows[0].keys()) if rows else [],
        "template": template,
    }

@app.post("/api/price-compare/import")
async def execute_import(file: UploadFile = File(...), template_name: str = Form(""), skip_rembg: str = Form("0")):
    """执行入库：上传 → 映射 → 逐条 ingest_product（后台线程 + 进度轮询）"""
    file_bytes = await file.read()
    rows = _parse_upload(file_bytes, file.filename or "data.csv")

    if not rows:
        return {"ok": False, "error": "文件无数据"}

    template = None
    if template_name:
        tpl_path = os.path.join(TEMPLATES_DIR, f"{template_name}.json")
        if os.path.exists(tpl_path):
            with open(tpl_path, encoding="utf-8") as fp:
                template = json.load(fp)

    job_id = uuid.uuid4().hex[:8]
    _import_jobs[job_id] = {
        "status": "running",
        "progress": {"done": 0, "total": len(rows)},
        "result": None,
    }

    threading.Thread(
        target=_run_import_job,
        args=(job_id, rows, template, template_name, skip_rembg == "1"),
        daemon=True,
    ).start()

    return {"ok": True, "job_id": job_id, "total": len(rows)}

# ── 平台自动识别 ──
PLATFORM_DOMAINS = {
    "tmall":       ["tmall.com", "detail.tmall.com", "list.tmall.com"],
    "taobao":      ["taobao.com", "item.taobao.com", "h5.m.taobao.com"],
    "jd":          ["jd.com", "item.jd.com", "i.jd.com"],
    "pdd":         ["pinduoduo.com", "yangkeduo.com", "mobile.yangkeduo.com"],
    "1688":        ["1688.com", "detail.1688.com"],
    "suning":      ["suning.com", "product.suning.com"],
    "vip":         ["vip.com", "detail.vip.com"],
    "amazon":      ["amazon.cn", "amazon.com", "www.amazon.com"],
    "douyin":      ["douyin.com", "haohuo.jinritemai.com"],
    "kuaishou":    ["kuaishou.com"],
    "dangdang":    ["dangdang.com", "product.dangdang.com"],
    "gome":        ["gome.com.cn", "item.gome.com.cn"],
    "meituan":     ["meituan.com", "i.meituan.com"],
    "xiaohongshu": ["xiaohongshu.com", "xhslink.com"],
    "yanxuan":     ["you.163.com", "yanxuan.com"],
    "weidian":     ["weidian.com", "shop.v.weidian.com"],
}

def detect_platform_from_url(url: str) -> str | None:
    """从商品 URL 中提取电商平台标识，失败返回 None。
    支持完整 URL、scheme-relative URL（//开头）、以及追踪/广告域名。"""
    if not url:
        return None
    try:
        from urllib.parse import urlparse
        # scheme-relative URL（//item.taobao.com/...）→ 补 https:
        if url.startswith("//"):
            url = "https:" + url
        parsed = urlparse(url)
        host = parsed.hostname or ""
        host_lower = host.lower().lstrip("www.")
        # 去掉移动端前缀
        host_lower = host_lower.removeprefix("m.").removeprefix("h5.").removeprefix("mobile.")
    except Exception:
        return None

    # 追踪/广告域名快速映射（比通用后缀匹配更优先）
    TRACKING_ALIASES = {
        "taobao": ["click.simba.taobao.com", "redirect.simba.taobao.com",
                     "s.click.taobao.com", "uland.taobao.com"],
        "tmall":  ["click.simba.tmall.com", "s.click.tmall.com"],
        "1688":   ["dj.1688.com", "s.click.1688.com", "ad.1688.com"],
    }
    for platform, aliases in TRACKING_ALIASES.items():
        if host_lower in aliases:
            return platform

    for platform, domains in PLATFORM_DOMAINS.items():
        for domain in domains:
            if host_lower == domain or host_lower.endswith("." + domain):
                return platform
    return None

def _run_import_job(job_id: str, rows: list, template: dict | None, template_name: str, skip_rembg: bool = False):
    """在后台线程中执行逐条入库"""
    job = _import_jobs[job_id]
    platform = "unknown"
    if template:
        platform = template.get("platform", "unknown")

    success = 0
    failed = 0
    missing_title = 0
    missing_image = 0
    missing_both = 0
    sample_missing = None
    errors = []

    col_names = list(rows[0].keys()) if rows else []

    start_time = time.time()

    # ── 第一阶段：解析每行，提取字段 ──
    parsed_rows = []
    for i, row in enumerate(rows):
        if template:
            mapped = {}
            for target_field, source_col in template.get("mappings", {}).items():
                mapped[target_field] = str(row.get(source_col, "")).strip()
        else:
            mapped = {k: str(v).strip() for k, v in row.items()}

        title = mapped.get("title", "") or mapped.get("标题", "") or mapped.get("商品标题", "") or mapped.get("产品标题", "") or mapped.get("名称", "") or mapped.get("商品名称", "")
        price_str = mapped.get("price") or mapped.get("价格") or mapped.get("售价") or mapped.get("单价") or "0"
        main_image_url = mapped.get("image_url", "") or mapped.get("main_image_url", "") or mapped.get("图片链接", "") or mapped.get("图片地址", "") or mapped.get("图片", "") or mapped.get("主图", "") or mapped.get("商品图片", "") or mapped.get("商品主图", "")
        url = mapped.get("url", "") or mapped.get("platform_url", "") or mapped.get("商品链接", "") or mapped.get("详情链接", "") or mapped.get("链接", "") or mapped.get("商品地址", "") or mapped.get("商品URL", "")
        description = mapped.get("description", "") or mapped.get("详情介绍", "") or mapped.get("描述", "") or mapped.get("商品描述", "") or mapped.get("详情", "")
        original_price_str = mapped.get("original_price", "") or mapped.get("原价", "") or mapped.get("划线价", "") or mapped.get("市场价", "")
        image_urls_raw = mapped.get("image_urls", "") or mapped.get("多图", "") or mapped.get("图片列表", "") or mapped.get("附图", "") or mapped.get("附图链接", "")
        image_urls = []
        if image_urls_raw:
            import re as _re
            _sep_pat = r"[,，;\n\r]+"
            image_urls = [u.strip() for u in _re.split(_sep_pat, image_urls_raw) if u.strip()]
        if not row_platform or row_platform == "unknown":
            detected = detect_platform_from_url(url)
            if detected:
                row_platform = detected
        shop_name = mapped.get("shop_name", "") or mapped.get("店铺", "") or mapped.get("店铺名", "") or mapped.get("店铺名称", "") or mapped.get("商家", "") or mapped.get("卖家", "")
        stock_info = mapped.get("stock_info", "") or mapped.get("库存", "") or mapped.get("库存/销量", "")
        shipping = mapped.get("shipping", "") or mapped.get("运费", "") or mapped.get("运费/物流", "")
        # SKU 原始值（暂存，价格算出来后再解析）
        sku_color = mapped.get("sku_color", "") or mapped.get("color", "") or mapped.get("颜色", "") or mapped.get("颜色分类", "")
        sku_size = mapped.get("sku_size", "") or mapped.get("size", "") or mapped.get("尺寸", "") or mapped.get("尺码", "")
        skus_raw = mapped.get("skus", "") or mapped.get("SKU", "") or mapped.get("sku", "")

        if not title or not main_image_url:
            parsed_rows.append(None)  # 标记为无效行
            continue

        try:
            price = float(_re.sub(r"[^\d.]", "", str(price_str))) if price_str else 0
        except Exception:
            price = 0

        # ── 解析 SKU（简单格式: 颜色|尺寸|价格|图索引;颜色|尺寸|价格|图索引;...）──
        parsed_skus = []
        if skus_raw:
            for seg in skus_raw.split(";"):
                seg = seg.strip()
                if not seg:
                    continue
                parts = [p.strip() for p in seg.split("|")]
                if not parts or not parts[0]:
                    continue
                sku_obj = {"color": parts[0], "size": parts[1] if len(parts) > 1 else "",
                          "price": float(parts[2]) if len(parts) > 2 and parts[2] else price,
                          "images": []}
                if len(parts) > 3 and parts[3]:
                    try:
                        sku_obj["images"] = [int(x.strip()) for x in parts[3].split(",")]
                    except ValueError:
                        sku_obj["images"] = [0]
                parsed_skus.append(sku_obj)
        # 兼容旧格式：独立颜色/尺寸列 → 整行作为一个 SKU
        if not parsed_skus and (sku_color or sku_size):
            parsed_skus = [{"color": sku_color, "size": sku_size, "price": price, "images": []}]

        parsed_rows.append({
            "title": title, "price": price, "main_image_url": main_image_url,
            "url": url, "platform": row_platform, "shop_name": shop_name,
            "description": description, "original_price_str": original_price_str,
            "image_urls": image_urls, "stock_info": stock_info, "shipping": shipping,
            "skus": parsed_skus if parsed_skus else [{"color": "", "size": "", "price": price, "images": []}],
        })

    # ── 第二阶段：按 (title, url) 分组合并 ──
    groups = {}  # key → {info, skus, all_images}
    for i, pr in enumerate(parsed_rows):
        if pr is None:
            continue
        key = (pr["title"], pr["url"] or pr["main_image_url"])
        if key not in groups:
            groups[key] = {
                "info": pr,
                "skus": [],
                "all_images": [pr["main_image_url"]] + pr["image_urls"],
                "image_map": {},  # url → index in all_images
            }
            groups[key]["image_map"][pr["main_image_url"]] = 0
            for idx, img_url in enumerate(pr["image_urls"]):
                groups[key]["image_map"][img_url] = idx + 1
        else:
            g = groups[key]
            # 合并图片（去重）
            for img_url in [pr["main_image_url"]] + pr["image_urls"]:
                if img_url and img_url not in g["image_map"]:
                    g["image_map"][img_url] = len(g["all_images"])
                    g["all_images"].append(img_url)
        # 添加 SKU（从新格式 or 兼容旧单 SKU）
        for s in pr.get("skus", []):
            sku_images = s.get("images", []) if s.get("images") else []
            groups[key]["skus"].append({
                "color": s.get("color", ""),
                "size": s.get("size", ""),
                "price": s.get("price", pr["price"]),
                "images": sku_images,
            })

    # ── 第三阶段：逐组入库 ──
    total = len(parsed_rows)
    done = 0
    success = 0
    failed = 0
    missing_title = sum(1 for p in parsed_rows if p is None and not any("title" in str(e) for e in []))  # will be refined
    missing_image = 0
    missing_both = 0
    sample_missing = None
    errors = []

    # 重新统计失败
    _none_count = sum(1 for p in parsed_rows if p is None)
    failed += _none_count
    missing_both += _none_count  # 全部标记为缺标题+缺图

    group_idx = 0
    for key, g in groups.items():
        group_idx += 1
        pr = g["info"]
        merged_skus = g["skus"]
        main_image_url = g["all_images"][0] if g["all_images"] else pr["main_image_url"]
        extra_images = g["all_images"][1:] if len(g["all_images"]) > 1 else []

        # 更新进度
        elapsed = time.time() - start_time
        job["progress"] = {
            "done": done,
            "total": total,
            "current": pr["title"][:40],
            "elapsed_s": round(elapsed, 1),
            "eta_s": 0,
            "step": f"入库中 {group_idx}/{len(groups)}",
        }

        # 如果只有一条 SKU 且无颜色尺寸，保持无 SKU 状态（向后兼容）
        final_skus = merged_skus if (len(merged_skus) > 1 or merged_skus[0].get("color") or merged_skus[0].get("size")) else []

        try:
            def img_progress(cur, total_imgs, step_label):
                job["progress"] = {
                    "done": done,
                    "total": total,
                    "current": pr["title"][:40],
                    "elapsed_s": round(time.time() - start_time, 1),
                    "eta_s": 0,
                    "step": f"{step_label}",
                    "img_cur": cur + 1,
                    "img_total": total_imgs,
                }

            pid = product_pipeline.ingest_product(
                platform=pr["platform"],
                url=pr["url"],
                title=pr["title"],
                price=pr["price"],
                shop_name=pr["shop_name"],
                main_image_url=main_image_url,
                attrs={},
                description=pr["description"],
                image_urls=extra_images,
                stock_info=pr["stock_info"],
                shipping=pr["shipping"],
                progress_callback=img_progress,
            )
            if pid:
                # 写入 skus JSON
                if final_skus:
                    db.update_product(pid, skus=final_skus)
                # 写入原价
                if pr["original_price_str"]:
                    try:
                        orig_price = float(_re.sub(r"[^\d.]", "", str(pr["original_price_str"])))
                        if orig_price > 0:
                            db.update_product(pid, original_price=orig_price)
                    except Exception:
                        pass
                success += 1
                done += len(merged_skus)  # 进展按 SKU 数计
            else:
                failed += 1
        except Exception as e:
            failed += 1
            if len(errors) < 10:
                errors.append(str(e))

    # 收尾统计

    job["status"] = "done"
    job["result"] = {
        "ok": True,
        "total": len(rows),
        "success": success,
        "failed": failed,
        "errors": errors[:10],
        "required_fields": ["title (标题)", "main_image_url (图片链接)"],
        "failure_detail": {
            "missing_title": missing_title,
            "missing_image": missing_image,
            "missing_both": missing_both,
        },
        "file_columns": col_names,
        "sample": sample_missing,
    }

@app.get("/api/price-compare/import-status/{job_id}")
async def get_import_status(job_id: str):
    """轮询入库进度"""
    job = _import_jobs.get(job_id)
    if not job:
        return {"ok": False, "error": "job not found"}
    return {
        "ok": True,
        "status": job["status"],
        "progress": job["progress"],
        "result": job["result"],
    }

# ═══════════════════════════════════════════
# 去背景预览 + 以图搜图 Demo
# ═══════════════════════════════════════════

import base64 as _base64

@app.post("/api/bg-remove/preview")
async def bg_remove_preview(file: UploadFile = File(...)):
    """上传图片 → 返回原图 + 去背景图的 base64"""
    try:
        data = await file.read()
        from PIL import Image
        from io import BytesIO
        import embedding

        # 原图
        orig_img = Image.open(BytesIO(data)).convert("RGB")
        orig_buf = BytesIO()
        orig_img.save(orig_buf, "JPEG", quality=85)
        orig_b64 = _base64.b64encode(orig_buf.getvalue()).decode()

        # 去背景
        processed = embedding.preprocess_image(orig_img)
        proc_buf = BytesIO()
        processed.save(proc_buf, "JPEG", quality=85)
        proc_b64 = _base64.b64encode(proc_buf.getvalue()).decode()

        return {
            "ok": True,
            "original": f"data:image/jpeg;base64,{orig_b64}",
            "processed": f"data:image/jpeg;base64,{proc_b64}",
        }
    except Exception as e:
        import traceback
        logger.error(f"[bg-remove/preview] {type(e).__name__}: {e}")
        logger.error(traceback.format_exc())
        return {"ok": False, "error": f"{type(e).__name__}: {e}"}

@app.post("/api/bg-remove/search")
async def bg_remove_search(file: UploadFile = File(...), top_k: int = 20, platform: str = ""):
    """上传图片 → 去背景 → 向量搜索 → 返回匹配商品"""
    try:
        data = await file.read()
        results = product_pipeline.search_by_image(
            image_data=data, top_k=top_k,
            platform=platform if platform else None
        )
        return {"ok": True, "results": results, "count": len(results)}
    except Exception as e:
        import traceback
        logger.error(f"[bg-remove/search] {type(e).__name__}: {e}")
        logger.error(traceback.format_exc())
        return {"ok": False, "error": f"{type(e).__name__}: {e}", "results": [], "count": 0}

# ── 去背景演示页面 ──
@app.get("/bg-demo")
@app.get("/bg-demo/")
async def bg_demo_page():
    """返回去背景 + 以图搜图演示页面"""
    demo_path = os.path.join(os.path.dirname(__file__), "bg_remove_demo.html")
    from fastapi.responses import HTMLResponse
    with open(demo_path, "r", encoding="utf-8") as f:
        return HTMLResponse(content=f.read())

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=19527, log_level="warning")
